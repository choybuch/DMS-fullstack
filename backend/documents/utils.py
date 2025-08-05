# backend/documents/utils.py
import fitz  # PyMuPDF
import docx
from openpyxl import load_workbook
import re
import tempfile
from django.core.files.storage import default_storage

def clean_text(text):
    """Clean and normalize extracted text"""
    if not text:
        return ""
    text = re.sub(r'\s+', ' ', text)  # Remove extra whitespace
    text = re.sub(r'[^\w\s]', '', text)  # Remove punctuation
    return text.lower()  # Case normalization

def extract_pdf_text(file):
    """Extract text from PDF files"""
    text = ""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
        for chunk in file.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name
    
    try:
        with fitz.open(tmp_path) as doc:
            text = "".join(page.get_text() for page in doc)
    finally:
        import os
        os.unlink(tmp_path)
    return clean_text(text)

def extract_docx_text(file):
    """Extract text from Word documents"""
    text = ""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as tmp:
        for chunk in file.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name
    
    try:
        doc = docx.Document(tmp_path)
        text = "\n".join(para.text for para in doc.paragraphs)
    finally:
        import os
        os.unlink(tmp_path)
    return clean_text(text)

def extract_excel_text(file):
    """Extract text from Excel files"""
    text_content = []
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        for chunk in file.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name
    
    try:
        wb = load_workbook(filename=tmp_path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                text_content.append(" ".join(
                    str(cell.value) for cell in row if cell.value
                ))
    finally:
        import os
        os.unlink(tmp_path)
    return clean_text("\n".join(text_content))